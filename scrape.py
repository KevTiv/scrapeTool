import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill



def scrape_legrand_websites(legrand_input, first_url):
    ############################################################################
    #  chrome driver
    # browser = webdriver.Chrome()
    # web driver path /Users/kevintivert/Documents/web_driver/chromedriver
    PATH = r'/usr/local/bin/chromedriver'

    is_product_found = False
    # chrome driver load
    driver = webdriver.Chrome(executable_path=PATH)
    ############################################################################
    # open excel file #
    legrand_excel = load_workbook('legrandParseOutput.xlsx')
    parse_file = legrand_excel.get_sheet_by_name('www.legrand.fr pro ')
    # color for cell with no value
    red_fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
    ############################################################################

    driver.execute_script("window.open('');")# Switch to the new window and open URL B
    driver.switch_to.window(driver.window_handles[1])

    for index in range( len(legrand_input) ):
        # driver.execute_script("window.open('');")# Switch to the new window and open URL B
        # driver.switch_to.window(driver.window_handles[1])
        driver.get(first_url + legrand_input[index])

        # Check if accept cookie pop up then click accept
        try:
            accept_button = driver.find_element_by_id('popin_tc_privacy_button')
            time.sleep(1)
            accept_button.click()
        except NoSuchElementException:

            # check_if_prod_is_found = driver.find_element_by_xpath('//*[@id="results"]/div/div/div/p') || span X resultats trouves //*[@id="results"]/div/div/div/p/span
            # # is_product_found = check_exists_by_xpath(check_if_prod_is_found)
            # is_product_found = check_if_prod_is_found.is_displayed()
            try:
                check_if_prod_is_not_found = driver.find_element_by_xpath('//*[@id="results"]/div/div/div/p/span')

                parse_file['A'+str(index + 1)] = legrand_input[index]
                parse_file['B'+str(index + 1)].fill = red_fill
                parse_file['C'+str(index + 1)].fill = red_fill
                parse_file['D'+str(index + 1)].fill = red_fill
            except NoSuchElementException:
                # if( driver.find_element_by_xpath('//*[@id="results"]/div/div/div/p').is_displayed() == False ): 
                    # find needed elements
                    img_element = driver.find_elements_by_xpath('/html/body/div[4]/div/div[1]/div/div[2]/div/div/div[1]/div[2]/div/div/div/div[2]/div/div/div/div/a/img')
                    caract_prd = driver.find_elements_by_xpath('//*[@id="caracteristiques-produit"]/div/div/div/div/div/div[1]/div/div[1]')
                    caract_general = driver.find_elements_by_xpath('//*[@id="caracteristiques-produit"]/div/div/div/div/div/div[2]/div/div')
                    # Write scraped data in an excel file
                    parse_file['A'+str(index + 1)] = legrand_input[index]

                    for my_prd_caract in caract_prd:
                        # print(my_prd_caract.text)
                        parse_file['B'+str(index + 1)] = my_prd_caract.text

                    time.sleep(1)
                    for my_gnrl_caract in caract_general:
                        # print(my_gnrl_caract.text)
                        parse_file['C'+str(index + 1)] = my_gnrl_caract.text

                    time.sleep(1)
                    for my_src in img_element:
                        # added index 0 because two url would  be written
                        # print(my_src[0].get_attribute("src"))
                        parse_file['D'+str(index + 1)] = my_src.get_attribute("src")

            legrand_excel.save('legrandParseOutput.xlsx')

                # driver.close()
            # else:
            #     #paint excel row red
            #     parse_file['A'+str(index+ 2)] = legrand_input[index]
            #     parse_file['B'+str(index+ 2)].fill = red_fill
            #     parse_file['C'+str(index+ 2)].fill = red_fill
            #     parse_file['D'+str(index+ 2)].fill = red_fill

            # driver.close()
     ############################################################################
    # save excel file #
    legrand_excel.save('legrandParseOutput.xlsx')
    driver.quit()
    ############################################################################


# def check_exists_by_xpath(is_prd_found):
#     try:
#         webdriver.find_element_by_xpath(is_prd_found)
#     except NoSuchElementException:
#         return False
#     return True

def legrand_ref_input_function():
    # open legrand reference input text file &
    # insert reference in an array
    input_file = open("legrand_ref_input.txt", "r")
    ref_input_array = input_file.read().splitlines()
    input_file.close()
    return ref_input_array


def main():
     # First legrand URL
    first_legrand_url= 'https://www.legrand.fr/pro/recherche?s='
    # Second legrand URL
    second_legrand_url = 'https://www.export.legrand.com/search/fr?words='

    legrand_ref_inputs = legrand_ref_input_function()

    scrape_legrand_websites(legrand_ref_inputs, first_legrand_url)

if __name__ == "__main__":
    main()
